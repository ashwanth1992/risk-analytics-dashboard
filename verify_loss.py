"""
verify_loss.py — Independent verification of D1-based business loss
Reads dashboard_data.json + the exported Excel, cross-checks every row.

Usage:
    python verify_loss.py                                    # defaults
    python verify_loss.py --excel rde_summary_2026-05-28.xlsx
    python verify_loss.py --data dashboard_data.json
"""

import json
import argparse
from pathlib import Path
from collections import defaultdict

def load_data(data_path):
    with open(data_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def fmt_pin(v):
    """Match engine_data.py _fmt_pin: zero-padded 6-digit string."""
    try:
        return str(int(float(v))).zfill(6)
    except:
        return str(v).strip()

def classify_rates(cat_rates, thr):
    """Mirror JS classifyRates: returns (cls, trigger)."""
    lr = cat_rates.get('Low')
    mr = cat_rates.get('Medium')
    hr = cat_rates.get('High')
    vr = cat_rates.get('Very High')
    
    if lr is not None and lr > thr['l']:
        return 'red_nonop', f"Low {lr*100:.2f}% > {thr['l']*100:.2f}%"
    if mr is not None and mr > thr['m']:
        return 'red_op', f"Medium {mr*100:.2f}% > {thr['m']*100:.2f}%"
    if hr is not None and hr > thr['h']:
        return 'red_op', f"High {hr*100:.2f}% > {thr['h']*100:.2f}%"
    if vr is not None and vr > thr['vh']:
        return 'red_op', f"VH {vr*100:.2f}% > {thr['vh']*100:.2f}%"
    return 'green', '—'

def get_rates_from_cats(cats):
    """Extract per-tier rates from R_DATA district cats."""
    out = {}
    for t in ['Low', 'Medium', 'High', 'Very High']:
        c = cats.get(t, {})
        total = c.get('total', 0)
        bad = c.get('bad', 0)
        out[t] = bad / total if total > 0 else None
    return out

def main():
    parser = argparse.ArgumentParser(description="Verify D1 business loss independently")
    parser.add_argument("--data", type=Path, default=Path("dashboard_data.json"))
    parser.add_argument("--excel", type=Path, default=None, help="Optional: exported Excel to cross-check")
    parser.add_argument("--sl", type=float, default=7.5, help="Low threshold %%")
    parser.add_argument("--sm", type=float, default=10.5, help="Medium threshold %%")
    parser.add_argument("--sh", type=float, default=10.0, help="High threshold %%")
    parser.add_argument("--svh", type=float, default=10.0, help="Very High threshold %%")
    args = parser.parse_args()
    
    thr = {'l': args.sl/100, 'm': args.sm/100, 'h': args.sh/100, 'vh': args.svh/100}
    
    print(f"Loading {args.data}...")
    data = load_data(args.data)
    
    r_data = data.get('REGION_DATA', {})
    lead_array = data.get('LEAD_ARRAY', [])
    d1_vol = data.get('D1_PINCODE_VOLUME', {})
    ats_data = data.get('ATS_DATA_JSON', {})
    window = data.get('WINDOW_CONFIG', {})
    risk_months = window.get('risk', {}).get('months', 7)
    
    print(f"LEAD_ARRAY: {len(lead_array):,} leads")
    print(f"D1_PINCODE_VOLUME: {len(d1_vol):,} pincodes")
    print(f"Risk months: {risk_months}")
    print(f"Thresholds: L={args.sl}% M={args.sm}% H={args.sh}% VH={args.svh}%")
    print()
    
    # ATS per tier
    ats = {}
    for t in ['Low', 'Medium', 'High', 'Very High']:
        ats[t] = ats_data.get(t, {}).get('ats', 0)
        print(f"  ATS[{t}] = ₹{ats[t]:,.0f}")
    print()
    
    # ── Step 1: Classify every district ──────────────────────────────────
    dist_cls = {}  # (state, district) → (cls, trigger)
    for sn, sd in r_data.items():
        for dn, dd in sd.get('districts', {}).items():
            cats = dd.get('cats', {})
            rates = get_rates_from_cats(cats)
            cls, trigger = classify_rates(rates, thr)
            dist_cls[(sn, dn)] = (cls, trigger, rates)
    
    red_nonop = [(k,v) for k,v in dist_cls.items() if v[0] == 'red_nonop']
    red_op = [(k,v) for k,v in dist_cls.items() if v[0] == 'red_op']
    green = [(k,v) for k,v in dist_cls.items() if v[0] == 'green']
    print(f"District classification: {len(red_nonop)} red_nonop, {len(red_op)} red_op, {len(green)} green")
    print()
    
    # ── Step 2: For each district, compute D1-based loss ─────────────────
    # Group leads by (state, district, pin, tier) 
    lead_groups = defaultdict(lambda: {'count': 0, 'bad': 0})
    for l in lead_array:
        st = l.get('cs') or l.get('s', '')
        dt = l.get('cd') or l.get('d', '')
        pin = str(l.get('pin', ''))
        tier = l.get('r', '')
        if not st or not dt or not tier:
            continue
        lead_groups[(st, dt, pin, tier)]['count'] += 1
        lead_groups[(st, dt, pin, tier)]['bad'] += l.get('b', 0)
    
    # For each stopped district, compute loss from D1 volumes
    d1_processed = set()
    d1_zero = set()
    total_loss_d1 = 0
    total_loss_hist = 0
    dist_loss = defaultdict(lambda: {'d1': 0, 'hist': 0, 'leads': 0, 'cls': '', 'pins_with_d1': 0, 'pins_without_d1': 0})
    
    for (st, dt, pin, tier), grp in lead_groups.items():
        cls_info = dist_cls.get((st, dt))
        if not cls_info:
            continue
        cls, trigger, rates = cls_info
        
        # Determine if this lead is dropped
        is_high_vh = tier in ('High', 'Very High')
        dropped = False
        if cls == 'red_nonop':
            dropped = True
        elif cls == 'red_op':
            if is_high_vh:
                dropped = True
        
        if not dropped:
            continue
        
        dk = f"{dt} ({st})"
        dist_loss[dk]['cls'] = cls
        dist_loss[dk]['leads'] += grp['count']
        
        # D1 volume lookup
        d1pin = fmt_pin(pin)
        pin_key = f"{d1pin}|{tier}"
        
        # D1 loss (deduplicated per pin-tier)
        if pin_key not in d1_processed:
            d1_processed.add(pin_key)
            d1v = d1_vol.get(d1pin, {}).get(tier)
            if d1v is not None and d1v > 0:
                loss_d1 = d1v * ats[tier]  # already monthly
                total_loss_d1 += loss_d1
                dist_loss[dk]['d1'] += loss_d1
                dist_loss[dk]['pins_with_d1'] += 1
            else:
                d1_zero.add(pin_key)
                dist_loss[dk]['pins_without_d1'] += 1
        
        # Historical loss (for comparison)
        total_loss_hist += grp['count'] * ats[tier]
        dist_loss[dk]['hist'] += grp['count'] * ats[tier]
    
    # ── Results ──────────────────────────────────────────────────────────
    print("=" * 80)
    print(f"{'VERIFICATION RESULTS':^80}")
    print("=" * 80)
    print()
    print(f"D1 Monthly Loss:         ₹{total_loss_d1/100000:.2f} L")
    print(f"Historical Loss (÷{risk_months}mo): ₹{total_loss_hist/risk_months/100000:.2f} L")
    print(f"Zero-D1 pin-tier combos: {len(d1_zero)}")
    print(f"Pin-tier combos with D1: {len(d1_processed) - len(d1_zero)}")
    print()
    
    print(f"{'District':<35} {'Cls':<12} {'Leads':>6} {'D1 Loss/mo':>12} {'Hist Loss/mo':>12} {'D1 Pins':>8} {'No-D1':>6}")
    print("-" * 95)
    for dk, v in sorted(dist_loss.items(), key=lambda x: -x[1]['d1']):
        d1_mo = v['d1'] / 100000
        hist_mo = v['hist'] / risk_months / 100000
        print(f"  {dk:<33} {v['cls']:<12} {v['leads']:>6} ₹{d1_mo:>9.2f}L ₹{hist_mo:>9.2f}L {v['pins_with_d1']:>8} {v['pins_without_d1']:>6}")
    
    # ── Cross-check with Excel if provided ───────────────────────────────
    if args.excel and args.excel.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(args.excel)
            ws = wb['Tightening']
            headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
            
            loss_col = headers.index('Loss/mo (₹L)') + 1
            d1vol_col = headers.index('D1 Vol/mo') + 1
            pin_col = headers.index('Pincode') + 1
            tier_col = headers.index('Risk Tier') + 1
            dist_col = headers.index('District') + 1
            
            excel_total = sum(ws.cell(r, loss_col).value or 0 for r in range(2, ws.max_row+1))
            
            print()
            print(f"{'EXCEL CROSS-CHECK':^80}")
            print(f"Excel total loss:   ₹{excel_total:.2f}L")
            print(f"Python D1 loss:     ₹{total_loss_d1/100000:.2f}L")
            diff = abs(excel_total - total_loss_d1/100000)
            print(f"Difference:         ₹{diff:.2f}L {'✓ OK (rounding)' if diff < 5 else '⚠ CHECK'}")
            
            # Spot-check 5 rows
            print()
            print("Spot-check (5 D1-covered rows):")
            checked = 0
            for r in range(2, ws.max_row+1):
                d1v = ws.cell(r, d1vol_col).value
                if d1v and d1v > 0 and checked < 5:
                    pin = str(ws.cell(r, pin_col).value)
                    tier = ws.cell(r, tier_col).value
                    actual_loss = ws.cell(r, loss_col).value or 0
                    expected = d1v * ats.get(tier, 0) / 100000
                    match = abs(actual_loss - expected) < 0.02
                    district = ws.cell(r, dist_col).value
                    print(f"  {'✓' if match else '✗'} {district}/{pin} {tier}: D1={d1v} × ATS=₹{ats.get(tier,0):,.0f} = ₹{expected:.2f}L (Excel: ₹{actual_loss:.2f}L)")
                    checked += 1
        except Exception as e:
            print(f"Excel cross-check failed: {e}")
    
    print()
    print("Done. If D1 Monthly Loss matches the dashboard, the calculation is correct.")

if __name__ == "__main__":
    main()